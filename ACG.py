from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import cv2 as cv 
import openpyxl
import pandas as pd
import os
import numpy as np




def browseFiles_files(): 
    filename = filedialog.askopenfilename(initialdir = "/",  title = "Select a File", filetypes = (("Excel files", 
                                                        ".xlsx"), ("all files", "."))) 
       
    # Change label contents 
    file_locationField.configure(text=filename)
    data_file = filename  
    #file name
    file = data_file[data_file.rfind("/")+1:]
    #data file location
    data_file = data_file[0:data_file.rfind("/")]
    #storing them to global variable
    Store_Location(data_file,file)
    
def Store_Location(filelocation,file):
    global location
    global data
    data = str()
    location = str()
    location = filelocation
    data = file
    
def browseFiles_img():

    filename = filedialog.askopenfilename(initialdir =  "/", title = "Select A File", filetypes =(("jpg files",
                                                              "*.jpg"),("all files","*.*")) )
    img_locationField.configure(text=filename)
    Store_certificate_sample(filename)
    

def Store_certificate_sample(file):
    global img_location
    img_location = str()
    
    img_location = file[file.rfind("/")+1:]
          
    
    
def generate_certificates():
    
    webinar_name = EventField.get()
    webinar_date = date_eventField.get()
    #certificate sample in templete_path
    template_path = img_location
    #location of certificate and excel file
    data_file = location
    #excel file in details path
    details_path = data
    #reading the excel file
    df = pd.read_excel(details_path)

    # Output Paths 
    output_path = data_file
    #making new folder named certificates
    new_folder_path = 'Certificates'
    #final directory
    final_path = os.path.join(output_path, new_folder_path)
    #creating the directory
    try:
        os.mkdir(final_path)
    except:
        pass
    finally:
        # Setting the font size and font 
        # colour 
        font_size = 0.9
        font_color = (0, 0, 00) 

        # Coordinates on the certificate where 
        # will be printing the name (set 
        # according to your own template) 
        coordinate_y_name = 30
        coordinate_x_name = 0

        coordinate_y_meetname = 5
        coordinate_x_meetname = 72

        coordinate_y_date = -15
        coordinate_x_date = -125
        date_font_size = 0.9
        # loading the details.xlsx workbook  
        # and grabbing the active sheet 
        obj = openpyxl.load_workbook(details_path) 
        sheet = obj.active 

        # printing for the first 10 names in the 
        # excel sheet 
        last_index=df['First Name'].size
        for i in range(2,last_index+1): 

            # grabs the row=i and column=1 cell  
            # that contains the name value of that 
            # cell is stored in the variable certi_name

            get_first_name = sheet.cell(row = i ,column = 1).value
            if get_first_name is None:
                get_first_name = " "
            get_last_name = sheet.cell(row = i ,column = 2).value
            if get_last_name is None:
                get_last_name = " "
            if (get_first_name == get_last_name and get_first_name == " "):
                continue
            certi_name = get_first_name.upper() +" "+ get_last_name.upper()
            approval_status= sheet.cell(row=i, column = 5).value
            if(approval_status == "approved"):
                # read the certificate template 
                img = cv.imread(template_path) 

                # choose the font from opencv 
                font = cv.FONT_HERSHEY_COMPLEX_SMALL          
                date_font = cv.FONT_HERSHEY_COMPLEX_SMALL
                # get the size of the name to be 
                # printed

                text_size = cv.getTextSize(certi_name, font, font_size, 10)[0]      

                # get the (x,y) coordinates where the 
                # name is to written on the template 
                # The function cv.putText accepts only 
                # integer arguments so convert it into 'int'. 
                text_x = (img.shape[1] - text_size[0]) / 2 + coordinate_x_name  
                text_y = (img.shape[0] + text_size[1]) / 2 - coordinate_y_name 
                text_x = int(text_x) 
                text_y = int(text_y) 
                cv.putText(img, certi_name, 
                          (text_x ,text_y ),  
                          font, 
                          font_size, 
                          font_color, 2) 
                
                text_size = cv.getTextSize(webinar_name, font, font_size, 7)[0]
                text_x = (img.shape[1] - text_size[0]) / 2 + coordinate_x_meetname  
                text_y = (img.shape[0] + text_size[1]) / 2 - coordinate_y_meetname 
                text_x = int(text_x) 
                text_y = int(text_y) 
                cv.putText(img, webinar_name, 
                          (text_x ,text_y ),  
                          font, 
                          font_size, 
                          font_color, 2) 
                
                text_size = cv.getTextSize(webinar_date, font, font_size, 10)[0]
                text_x = (img.shape[1] - text_size[0]) / 2  + coordinate_x_date  
                text_y = (img.shape[0] + text_size[1]) / 2  - coordinate_y_date 
                #text_x =  coordinate_x_date  
                #text_y =  coordinate_y_date 
                text_x = int(text_x) 
                text_y = int(text_y) 
                cv.putText(img, webinar_date, 
                          (text_x ,text_y ),  
                          date_font, 
                          date_font_size, 
                          font_color, 2) 

                # Output path along with the name of the 
                # certificate generated 

                certi_path = final_path +'/' + certi_name + '.png'

                # Save the certificate                       
                cv.imwrite(certi_path,img)




# Driver Code 
if __name__ == "__main__" : 
    
    global data_file 
    global cert_location 
    global certificate
    global EventField
    global date_eventField
    global location
    global data
    data = str()
    location = str()
    
    # Create a GUI window 
    gui = Tk() 
  
    # Set the background colour of GUI window   
    gui.configure(background = "paleturquoise2") 
  
    # set the name of tkinter GUI window  
    gui.title("CERTIFICATE GENERATOR" ) 
  
     # Set the configuration of GUI window 
    gui.geometry("400x330") 
   
    ATTENDX = Label(gui, text = "AUTOMATIC CERTIFICATE GENERATOR", fg = "white" , bg ="grey8", width = 30) 
  
    event_name = Label(gui, text = "EVENT NAME ", bg = "peachpuff" , fg="grey3" , width = 15) 
    
    date_event = Label(gui, text = "DATE OF EVENT ", bg = "peachpuff" , fg="grey3",width = 15) 
  
    file_location = Label(gui, text = "FILE LOCATION ",bg = "peachpuff" , fg="grey3",width = 15)
    
    img_location = Label(gui, text = "IMAGE LOCATION ",bg = "peachpuff" , fg="grey3" , width = 15) 
    
    Submit_button = Button(gui, text= "Generate Certificates",bg ="black",fg ='white', command = generate_certificates ,width = 50 )
     
     
  
    
    
    button_explore_file = Button(gui, text = "Browse Files", command = browseFiles_files, bg= "khaki1", fg= "black", width = 12 )
    button_exp = Button(gui, text = "Browse Image", command = browseFiles_img, bg= "khaki1", fg= "black" , width =12)
    
    
    # Create a text entry box for filling or typing the information.   
    EventField = Entry(gui, width =19) 
    date_eventField = Entry(gui,fg ="black", width =19) 
    file_locationField =  Label(gui, text = ".....",bg = "cyan2" , fg="black" , width =16)
    img_locationField =  Label(gui, text = ".....",bg = "cyan2" , fg="black" , width =16)
    
    
    def instruct():
        instructions = Tk()
        instructions.configure(background = "black")
        instructions.title("Requirement" )  
        instructions.geometry("800x200") 

        instruct1 = Label(instructions, text = "1. Excel file must have First name as heading in the first column and Last name in the second column of the first row. ",bg = "peachpuff" , fg="grey3" )
        instruct1.place(x=10,y=10)
        instruct2 = Label(instructions, text = "2.Ceritificate must be in JPG format and it must be stored along with the excel file in the same folder.",bg = "peachpuff" , fg="grey3")
        instruct2.place(x=10,y=35)
        steps = Label(instructions, text = "STEPS TO FOLLOW ",bg = "RED" , fg="white" )
        step = Label(instructions, text = "1.Enter the Event name and Event date accordingly.                           \n2.Browse files and select the excel file.                                                   \n3.Browse files and select the Certificate file according to your need.\n4.Click on Generate Certificates.                                                             ",bg = "peachpuff" , fg="grey3" )
        steps.place(x=350,y=60)
        step.place(x=10,y=85)  
        instructions.mainloop()
        
    
    
    
    
    button_requirements = Button(gui, text="Read Instructions",command= instruct, bg="coral1",fg ="black",width=50)
    
    
    
    
    button_generate = Button(gui, text= "Generate Certificates", command = generate_certificates, bg="coral1", fg ="black")
  
    # grid method is used for placing   
    # the widgets at respective positions   
    # in table like structure . 
    ATTENDX.place(x=100,y=10)
    event_name.place(x=20 , y=50)
    EventField.place(x=150 , y=50)
    date_event.place(x= 20, y = 100)
    date_eventField.place(x=150, y=100)
    file_location.place(x=20, y=150)
    file_locationField.place(x=150, y=150)
    button_explore_file.place(x=280, y=150)
    img_location.place(x=20, y=200)
    img_locationField.place(x=150, y=200)
    button_exp.place(x=280, y=200)
    Submit_button.place(x=15, y=275)
    button_requirements.place(x=15,y=245)
    
     
  
    # Start the GUI 
    gui.mainloop()
